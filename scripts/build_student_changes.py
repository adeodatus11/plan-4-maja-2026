#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def iso_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = clean(value)
    for pattern in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            pass
    raise ValueError(f"Nieprawidłowa data: {text}")


def split_branch(value: object) -> tuple[str, str]:
    parts = [clean(part) for part in clean(value).split("|")]
    return parts[0], parts[1] if len(parts) > 1 else ""


def period_number(value: object) -> int:
    match = re.match(r"\s*(\d+)", clean(value))
    if not match:
        raise ValueError(f"Nieprawidłowy numer lekcji: {value}")
    return int(match.group(1))


def descriptor(value: object) -> dict[str, object]:
    text = clean(value)
    match = re.match(
        r"(\d{2}\.\d{2}\.\d{4}),\s*(\d+),\s*[^,]+,\s*sala:\s*(.+)$",
        text,
        flags=re.I,
    )
    if not match:
        raise ValueError(f"Nieprawidłowy opis przeniesienia: {text}")
    return {
        "date": iso_date(match.group(1)),
        "period": int(match.group(2)),
        "room": clean(match.group(3)),
    }


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Brak arkusza '{sheet_name}' w {path.name}")
    values = workbook[sheet_name].iter_rows(values_only=True)
    headers = [clean(value) for value in next(values)]
    return [dict(zip(headers, row)) for row in values if any(value not in (None, "") for value in row)]


def build_changes(substitutions_path: Path, transfers_path: Path) -> dict[str, object]:
    substitutions = []
    for row in sheet_rows(substitutions_path, "Oddziały"):
        class_name, group_name = split_branch(row.get("Oddział"))
        raw_substitute = clean(row.get("Zastępca"))
        is_message = raw_substitute.casefold().startswith("uczniowie ") or "złączenie grup" in raw_substitute.casefold()
        substitutions.append({
            "date": iso_date(row.get("Dzień")),
            "period": period_number(row.get("Lekcja")),
            "className": class_name,
            "groupName": group_name,
            "type": "message" if is_message else "substitution",
            "message": raw_substitute,
            "room": "" if is_message else clean(row.get("Sala")),
        })

    transfers = []
    for row in sheet_rows(transfers_path, "Oddziały"):
        source = descriptor(row.get("Przeniesiono z"))
        target = descriptor(row.get("Przeniesiono na"))
        class_name, group_name = split_branch(row.get("Oddział"))
        if source["date"] != target["date"] or source["period"] != target["period"]:
            raise ValueError("Nakładka uczniowska obsługuje tylko zmianę sali w tym samym terminie.")
        transfers.append({
            "date": source["date"],
            "period": source["period"],
            "className": class_name,
            "groupName": group_name,
            "type": "room",
            "fromRoom": source["room"],
            "toRoom": target["room"],
        })

    payload = {"substitutions": substitutions, "transfers": transfers}
    forbidden = {"payment", "reason", "note", "absent", "uwagi", "płat", "plat"}
    serialized_keys = " ".join(
        key.casefold()
        for collection in payload.values()
        for item in collection
        for key in item
    )
    if any(term in serialized_keys for term in forbidden):
        raise ValueError("W danych uczniowskich wykryto niedozwolone pole.")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Buduje bezpieczną nakładkę zmian dla planu uczniowskiego.")
    parser.add_argument("substitutions", type=Path)
    parser.add_argument("transfers", type=Path)
    parser.add_argument("--output", type=Path, default=ROOT / "student-changes.json")
    args = parser.parse_args()

    payload = build_changes(args.substitutions.resolve(), args.transfers.resolve())
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Zapisano {args.output}: {len(payload['substitutions'])} zastępstw, {len(payload['transfers'])} zmian sali")


if __name__ == "__main__":
    main()
